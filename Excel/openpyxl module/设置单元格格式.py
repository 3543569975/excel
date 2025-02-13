
from openpyxl.styles import Font, colors,Color, Alignment,Border, Side ,PatternFill

# 创建一个工作簿
from openpyxl import workbook
wb = workbook.Workbook()
sheet1 = wb.worksheets[0]

# 设置字体
font = Font(name='Arial', size=12, bold=True, italic=True, underline='single', color=colors.RED)
#name='Arial'字体, size=12字号, bold=True加粗,italic=True倾斜,
# underline=True下划线('single', 'doubleAccounting', 'double', 'singleAccounting'),
# color=colors.RED

# 设置对齐方式
alignment = Alignment(horizontal='center', vertical='top')

#设置边框
border = Border(
    left=Side(style='thin', color=colors.RED),
    right = Side(style='thin', color=colors.RED),
    top = Side(style='thin', color='000000'),
    bottom = Side(style='thin', color='000000')
)

#设置背景色
fill = PatternFill(start_color ='FFFF00',end_color=colors.RED,fill_type='solid')#fill_type='solid'单色填充
#在使用ARGB颜色代码时，格式应该是AARRGGBB，其中AA是透明度，RR是红色值，GG是绿色值，BB是蓝色值。在上面的例子中，'FFFF0000'代表了红色。


# 应用格式到单元格
ls = [1,2,3,4,45,6,7]
for i in range(len(ls)):
    sheet1.cell(i+1,1).value=ls[i]
    sheet1.cell(i+1,1).font = font
    sheet1.cell(i+1,1).alignment = alignment
    sheet1.cell(i+1,1).border = border
    sheet1.cell(i+1,1).fill = fill

# 保存工作簿
wb.save('单元格格式.xlsx')