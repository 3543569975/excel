from openpyxl import workbook
wb = workbook.Workbook()
sheet = wb.worksheets[0]
sheet.cell(1,1).value='kaishi'
sheet.cell(1,2).value='开始'
sheet.cell(1,3).value='123'
sheet.cell(2,3).value=int('123')
wb.save('1.xlsx')