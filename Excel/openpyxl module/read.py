#f = open('C:\\Users\\l\\OneDrive\\桌面\\L\\pycharm\\pycharm练习\\Excel\city.xlsx',mode= 'rb')
from openpyxl import load_workbook

wb = load_workbook('city.xlsx')
v1 = wb.sheetnames
print(v1)
'''
AttributeError: 'Workbook' object has no attribute 'worknames'
sheet = wb.worknames('2')
cell = sheet.cell(1,1)
print(cell.value)

v1 = wb.worksheets
print(v1)
sheet = wb.worksheets[1]
cell = sheet.cell(5,5)
print(cell.value)
'''
sheet = wb.worksheets[1]
for i in range(1,10):
    cell = sheet.cell(i, 2)
    if '' == cell.value:
        print((sheet.cell(i,1)).value)