from openpyxl import load_workbook

from tqdm import tqdm
'''
import time
for i in tqdm(range(1,940)):
    cell = sheet.cell(i, 8)
    time.sleep(0.1)
'''
wb = load_workbook('【群星3.10.4】天灾宝典.xlsx')
v1 = wb.sheetnames
print(v1)
sheet = wb.worksheets[12]
ls = []
for i in range(1,941):
    cell = sheet.cell(i, 8)
    cell1 = sheet.cell(i, 7)
    if '¤' == cell.value and '︽' == cell1.value:
        ls.append((sheet.cell(i, 2)).value)


from openpyxl import workbook
wb = workbook.Workbook()
sheet1 = wb.worksheets[0]
for i in range(len(ls)):
    sheet1.cell(i+1,1).value=ls[i]

wb.save('测试excel读写.xlsx')