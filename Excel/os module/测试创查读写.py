import os

folder_path = os.path.join('C:\\Users\\l\\OneDrive\\桌面\\L\\pycharm\\pycharm练习\\Excel\\Excel\\os模块\\表格')
biaonames=[]
for name in os.listdir(folder_path):
    file_path = os.path.join(folder_path,name)
    biaonames.append(file_path)

from openpyxl import load_workbook

ls = []
for k in biaonames:
    wb = load_workbook(k)

    sheet = wb.worksheets[1]

    for i in range(1,5):
        cell = sheet.cell(i, 1)
        if '2' in  cell.value:
            ls.append((sheet.cell(i, 2)).value)

file = '测试创查读写' #（os.path.join('Excel','os模块')
os.makedirs(file)

from openpyxl import workbook

address = '测试创查读写/测试创查读写.xlsx'
wb = workbook.Workbook()
sheet1 = wb.worksheets[0]
for i in range(len(ls)):
    sheet1.cell(i+1,1).value=ls[i]

wb.save('测试创查读写/测试创查读写.xlsx')